from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side

# а) Прочитайте из трёх excel файлов (заранее создайте их, внутри 1111, 2222, 3333).
ws1 = load_workbook("sheet_01.xlsx")
ws2 = load_workbook("sheet_02.xlsx")
ws3 = load_workbook("sheet_03.xlsx")

wb1 = ws1.active
wb2 = ws2.active
wb3 = ws3.active

data = [
    [wb3["D1"].value, wb3["C1"].value, wb3["B1"].value, wb3["A1"].value],
    [wb1["D1"].value, wb1["C1"].value, wb1["B1"].value, wb1["A1"].value],
    [wb2["D1"].value, wb2["C1"].value, wb2["B1"].value, wb2["A1"].value]
]
# б) Отсортируйте полученную матрицу в порядке убывания.
data.sort()
data.reverse()

# в) Запишите это в один файл, изменив шрифт и обернув в границы.
res_ws = Workbook()
res_wb = res_ws.active

# Стили границ
medium = Side(border_style="medium")
border = Border(top=medium, bottom=medium, left=medium, right=medium)

# Стили текста
text_style = Font(italic=True)

counter_i = 0
for i in res_wb["A1":"D3"]:
    counter_j = 0

    for j in i:
        j.value = data[counter_i][counter_j]
        j.font = text_style
        j.border = border

        counter_j += 1
    counter_i += 1

res_ws.save("res_sheet.xlsx")
