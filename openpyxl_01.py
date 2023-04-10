from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("Лист 1")

ws["A4"] = 5
print(ws['A4'])
print(ws['A4'].value)
print(ws['A4'].column)
print(ws['A4'].row)

c = ws['A4']
c.value *= 2
print(c.value)
print(ws['A4'].value)

cell_range = ws['A1':'C2']
print(cell_range)

wb.save("test.xlsx")
